"""trim parse 命令 - 多表格文件解析"""

import re
import warnings
from pathlib import Path
from typing import Optional, Tuple, List
import pandas as pd
from .excel_reader import ExcelReader
from .csv_exporter import CSVExporter


def remove_all_whitespace(text: str) -> str:
    """去除所有形式的空白字符，包括空格、制表符、换行符、全角空格等
    
    Args:
        text: 输入字符串
    
    Returns:
        去除所有空白字符后的字符串
    """
    # 去除全角空格
    text = text.replace("　", "")
    # 去除所有空白字符（空格、制表符、换行符等）
    text = re.sub(r'\s+', '', text)
    return text


def parse_cell_range(range_str: str) -> Tuple[int, int, int, int]:
    """解析单元格范围字符串，如 "B1:H2"
    
    Args:
        range_str: 范围字符串，格式如 "B1:H2" 或 "A1"
    
    Returns:
        (start_col, start_row, end_col, end_row)
    """
    if ":" in range_str:
        start, end = range_str.split(":")
        start_col, start_row = parse_cell_address(start)
        end_col, end_row = parse_cell_address(end)
    else:
        start_col, start_row = parse_cell_address(range_str)
        end_col, end_row = start_col, start_row
    
    return start_col, start_row, end_col, end_row


def parse_cell_address(address: str) -> Tuple[int, int]:
    """解析单元格地址，如 "B1"
    
    Args:
        address: 单元格地址
    
    Returns:
        (column, row)
    """
    match = re.match(r"([A-Z]+)(\d+)", address.upper())
    if not match:
        raise ValueError(f"无效的单元格地址: {address}")
    
    col_str, row_str = match.groups()
    col = column_letter_to_number(col_str)
    row = int(row_str)
    
    return col, row


def column_letter_to_number(col: str) -> int:
    """将列字母转换为数字，如 A->1, B->2, AA->27"""
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def column_number_to_letter(num: int) -> str:
    """将数字转换为列字母，如 1->A, 2->B, 27->AA"""
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _get_merged_cell_value(ws, row: int, col: int, strip=True):
    """获取单元格值，处理合并单元格的情况
    
    Args:
        ws: worksheet对象
        row: 行号
        col: 列号
        strip: 是否去除前后空格
    
    Returns:
        单元格值，如果是合并单元格则返回合并区域的值
    """
    cell = ws.cell(row=row, column=col)
    
    # 检查是否在合并范围内
    for merged_range in ws.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1):
            if col in range(merged_range.min_col, merged_range.max_col + 1):
                # 返回合并区域左上角的单元格值
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                val = top_left_cell.value
                if val and strip:
                    val = str(val).strip()
                return val
    
    val = cell.value
    if val and strip:
        val = str(val).strip()
    return val


# 企业信息映射表（企业全称 -> (企业简称, 企业ID)）
ENTERPRISE_MAP = {
    "内蒙古包头鑫达黄金矿业有限责任公司": ("包头鑫达", "14000BTXD200403"),
    "广西凤山天承黄金矿业有限责任公司": ("凤山天承", "547600FSTC200801"),
    "中国黄金集团西和矿业有限公司": ("甘肃西和", "742100GSXH200801"),
    "中国黄金集团石湖矿业有限公司": ("石湖矿业", "50500HBSH200601"),
    "秦皇岛北戴河黄金宾馆有限公司": ("北戴河宾馆", "66100BDHBG199501"),
    "和布克赛尔蒙古自治县金洲矿业有限公司": ("和布克赛尔金洲", "834400HBKSEJZ200901"),
    "河北大白阳金矿有限公司": ("河北大白阳", "75100HBDBY199801"),
    "中国黄金集团河北有限公司": ("河北有限", "50000HBYX200801"),
    "河北中金黄金有限公司": ("河北中金", "50000HBZJ201001"),
    "吉木乃金源矿业有限公司": ("吉木乃金源", "836800JMNJY200501"),
    "河北峪耳崖黄金矿业有限责任公司": ("河北峪耳崖", "67600HBYEY199701"),
    "河北东梁黄金矿业有限责任公司": ("河北东梁", "67000HBDL200701"),
    "河北金厂峪矿业有限责任公司": ("河北金厂峪", "64300HBJCY200001"),
    "迁西县金顺商贸有限公司": ("迁西金顺", "64300QXJS200701"),
    "河南金源黄金矿业有限责任公司": ("河南金源", "471400HNJY200208"),
    "中国黄金河南有限公司": ("河南公司", "450000HNGS201208"),
    "河南黄金物资有限公司": ("河南物资", "450000HNWZ200601"),
    "河南秦岭黄金矿业有限责任公司": ("河南秦岭", "472500HNQL199701"),
    "河南金渠黄金股份有限公司": ("河南金渠", "472500HNJQ199807"),
    "河南省三门峡黄金工业学校": ("三门峡中金", "472000SMXZJ200604"),
    "黑龙江乌拉嘎黄金矿业有限责任公司": ("黑龙江乌拉嘎", "153200HLJWLG199708"),
    "中国黄金集团黑龙江有限公司": ("黑龙江公司", "150000HLJGS201201"),
    "黑龙江乌拉嘎金矿": ("乌拉嘎金矿", "153200WLGJK194901"),
    "黑龙江中金农业发展有限责任公司": ("中金农发", "150000ZJNF201301"),
    "湖北三鑫金铜股份有限公司": ("湖北三鑫", "435100HBSX199912"),
    "安徽太平矿业有限公司": ("安徽太平", "233100AHTP200508"),
    "湖北鸡笼山黄金矿业有限公司": ("湖北鸡笼山", "435200HBJLS199912"),
    "黄石金达矿业有限公司": ("黄石金达", "435000HSJD200201"),
    "四川通用投资有限公司": ("四川通用", "610000SCTY200903"),
    "中国黄金集团夹皮沟矿业有限公司": ("吉林夹皮沟", "132400JLJPG200508"),
    "中国黄金集团江西金山矿业有限公司": ("江西金山", "334200JXJS200201"),
    "莱州汇金矿业投资有限公司": ("莱州汇金", "261400LZHJ200704"),
    "辽宁中金黄金有限责任公司": ("辽宁中金", "122000LNZJ200501"),
    "凤城市瑞锦矿业有限公司": ("凤城瑞锦", "118100FCRJ201001"),
    "辽宁二道沟黄金矿业有限责任公司": ("辽宁二道沟", "122100LNEDG199901"),
    "辽宁金凤黄金矿业有限责任公司": ("辽宁金凤", "123000LNJF200011"),
    "辽宁排山楼黄金矿业有限责任公司": ("辽宁排山楼", "123100LNPSL199608"),
    "辽宁天择黄金矿业有限公司": ("辽宁天择", "122000LNTZ200701"),
    "辽宁省黄金物资有限公司": ("辽宁物资", "110000LNWZ199201"),
    "中国黄金集团辽宁有限公司": ("辽宁有限", "110000LNYX200712"),
    "凌源日兴矿业有限公司": ("凌源日兴", "122500LYRX200311"),
    "中国黄金集团辽宁有限公司海南分公司": ("辽宁有限海南", "572002LNYXHN201108"),
    "内蒙古金陶股份有限公司": ("内蒙金陶", "24300NMGJT199912"),
    "中国黄金集团内蒙古矿业有限公司": ("内蒙古矿业", "21300NMGKY200507"),
    "山东烟台鑫泰黄金矿业有限责任公司": ("山东鑫泰", "264000SDXT200511"),
    "烟台平泰安全技术服务有限公司": ("烟台平泰", "264000YTPT201701"),
    "陕西久盛矿业投资管理有限公司": ("陕西久盛", "710000SXJS200805"),
    "陕西鑫元科工贸股份有限公司": ("陕西鑫元", "710000SXXY200801"),
    "陕西太白黄金矿业有限责任公司": ("陕西太白", "721600SXTB200408"),
    "甘肃中金黄金矿业有限责任公司": ("甘肃中金", "746000GSZJ200601"),
    "中国黄金集团陕西有限公司": ("陕西公司", "710000SXGS200808"),
    "陕西略阳铧厂沟金矿有限公司": ("陕西铧厂沟", "724300SXHCG200601"),
    "陕西太白金矿": ("太白金矿", "721600TBJK198801"),
    "嵩县金牛有限责任公司": ("嵩县金牛", "471400SXJN199709"),
    "嵩县前河矿业有限责任公司": ("嵩县前河", "471400SXQH200211"),
    "苏尼特金曦黄金矿业有限责任公司": ("苏尼特金曦", "11200SNTJX200612"),
    "甘肃省天水李子金矿有限公司": ("天水李子", "741000TSLZ199501"),
    "潼关中金黄金矿业有限责任公司": ("潼关矿业", "714300TGKY200601"),
    "中国黄金集团新疆金滩矿业有限公司": ("新疆金滩", "839000XJJT200701"),
    "托里县金福黄金矿业有限责任公司": ("托里金福", "834500TLJF200501"),
    "云南黄金有限责任公司": ("云南公司", "650000YNHJ200108"),
    "西盟云天矿业有限公司": ("西盟云天", "665700XMYT201201"),
    "云南黄金有限责任公司镇沅分公司": ("云南镇沅", "666500YNZY200403"),
    "云南黄金有限责任公司新平分公司": ("云南新平", "653400YNXP200801"),
    "河南中原黄金冶炼厂有限责任公司": ("中原冶炼厂", "472000ZYYLC199510"),
    "江西三和金业有限公司": ("江西三和", "332000JXSH200405"),
    "辽宁天利金业有限责任公司": ("辽宁天利", "123000LNTL200601"),
    "辽宁新都黄金有限责任公司": ("辽宁新都", "122100LNXD199301"),
    "中金嵩县嵩原黄金冶炼有限责任公司": ("嵩原冶炼", "471400SYYL200412"),
    "潼关中金冶炼有限责任公司": ("潼关冶炼", "714300TGYL199711"),
    "西藏中金黄金冶炼有限公司": ("西藏中金", "850000XZZJ201301"),
    "河南中金中原新材料有限责任公司": ("中原新材料", "472000ZYXCL201101"),
    "三门峡中金矿业投资有限公司": ("三门峡中金", "472000SMXZJ200512"),
    "中国黄金集团科技有限公司": ("中金科技", "102600ZJKJ200504"),
}


# 括号内容到字母缩写的映射
BRACKET_CONTENT_MAP = {
    "本部": "BB",
    "管理": "GL",
    "差额": "CE",
    "合并": "HB",
    "小合并": "XHB",
    "小合并差额": "XHBCE",
    "（合并）": "HB",
    "（差额）": "CE",
    "（本部）": "BB",
    "（管理）": "GL",
    "（小合并）": "XHB",
    "（小合并差额）": "XHBCE",
}


def parse_sheet_name(sheet_name: str) -> Tuple[str, str]:
    """解析sheet名字，提取公司名称和括号内容
    
    Args:
        sheet_name: sheet名字，格式如 "公司名称(括号内容)" 或 "公司名称（括号内容）"
    
    Returns:
        (公司名称, 括号内容缩写)，括号内容默认为"BB"
    """
    # 匹配 名称(内容) 或 名称（内容） 格式
    match = re.match(r'^(.+?)（(.+?)）$', sheet_name.strip())
    if match:
        company = match.group(1).strip()
        bracket_content = match.group(2).strip()
    else:
        # 尝试匹配英文括号
        match = re.match(r'^(.+?)\(([^)]+)\)$', sheet_name.strip())
        if match:
            company = match.group(1).strip()
            bracket_content = match.group(2).strip()
        else:
            # 没有括号，整个名称为公司名称
            company = sheet_name.strip()
            bracket_content = "本部"
    
    # 转换为字母缩写
    bracket_code = BRACKET_CONTENT_MAP.get(bracket_content, "BB")
    return company, bracket_code


def parse_excel_with_axis(
    file_path: str,
    output_dir: str = ".",
    haxis: Optional[str] = None,
    vaxis: Optional[str] = None,
    merge: bool = False,
    timestamp: Optional[str] = None,
    name: Optional[str] = None,
) -> List[str]:
    """使用行列轴解析Excel文件
    
    Args:
        file_path: Excel文件路径
        output_dir: 输出目录
        haxis: 列标题范围，如 "B1:H2"
        vaxis: 行标题范围，如 "A2:B10"
        merge: 是否合并模式（所有sheet合并到一个文件）
        timestamp: 会计期间值，直接使用传入的字符串
        name: 数据集值，直接使用传入的字符串
    
    Returns:
        导出的CSV文件路径列表
    """
    reader = ExcelReader(file_path)
    exporter = CSVExporter(output_dir)
    
    try:
        output_files = []
        
        all_dfs = []
        
        for sheet_name in reader.get_sheet_names():
            ws = reader.workbook[sheet_name]
            
            # 解析轴范围
            h_start_col, h_start_row, h_end_col, h_end_row = (1, 1, 0, 0)
            v_start_col, v_start_row, v_end_col, v_end_row = (1, 1, 0, 0)
            
            if haxis:
                h_start_col, h_start_row, h_end_col, h_end_row = parse_cell_range(haxis)
            
            if vaxis:
                v_start_col, v_start_row, v_end_col, v_end_row = parse_cell_range(vaxis)
            
            # 确定数据区域
            if haxis and vaxis:
                # 两者都有：数据区域是行列标题范围的交叉区域
                data_start_row = v_start_row
                data_start_col = h_start_col
                data_end_row = v_end_row
                data_end_col = h_end_col
            elif haxis:
                # 只有列标题：数据在列标题下方，同一列范围内
                data_start_row = h_end_row + 1
                data_start_col = h_start_col
                data_end_row = ws.max_row
                data_end_col = h_end_col
            elif vaxis:
                # 只有行标题：数据在行标题右侧，同一行范围内
                data_start_row = v_start_row
                data_start_col = v_end_col + 1
                data_end_row = v_end_row
                data_end_col = ws.max_column
            else:
                # 都没有：读取整个sheet
                data_start_row = ws.min_row
                data_start_col = ws.min_column
                data_end_row = ws.max_row
                data_end_col = ws.max_column
            
            # 读取数据
            data = []
            for row in range(data_start_row, data_end_row + 1):
                row_data = []
                for col in range(data_start_col, data_end_col + 1):
                    val = _get_merged_cell_value(ws, row, col, strip=False)
                    # 非数字的单元格置空
                    if val is not None and not isinstance(val, (int, float)):
                        val = None
                    row_data.append(val)
                data.append(row_data)
            
            if not data:
                continue
            
            # 读取并合并列标题
            merged_col_headers = []
            if haxis:
                for col_idx in range(h_start_col, h_end_col + 1):
                    parts = []
                    for row_idx in range(h_start_row, h_end_row + 1):
                        val = _get_merged_cell_value(ws, row_idx, col_idx)
                        if val:
                            # 尝试转int（先转float再转int，处理58.0这样的格式）
                            try:
                                val = int(float(val))
                            except (ValueError, TypeError):
                                pass
                            val = remove_all_whitespace(str(val))
                            parts.append(val)
                    merged_col_headers.append("_".join(parts) if parts else f"Col_{col_idx}")
            else:
                from openpyxl.utils import get_column_letter
                for col in range(data_start_col, data_end_col + 1):
                    merged_col_headers.append(get_column_letter(col))
            
            # 读取并合并行标题
            if vaxis:
                row_headers = []
                for row_idx in range(v_start_row, v_end_row + 1):
                    parts = []
                    for col_idx in range(v_start_col, v_end_col + 1):
                        val = _get_merged_cell_value(ws, row_idx, col_idx)
                        if val:
                            # 尝试转int（先转float再转int，处理58.0这样的格式）
                            try:
                                val = int(float(val))
                            except (ValueError, TypeError):
                                pass
                            val = remove_all_whitespace(str(val))
                            parts.append(val)
                    row_headers.append("_".join(parts) if parts else f"Row_{row_idx}")
            else:
                row_headers = None
            
            # 创建DataFrame
            if merge:
                # 合并模式：每个sheet一行数据
                # 解析sheet名字
                company, bracket_content = parse_sheet_name(sheet_name)
                
                # 根据企业全称查找简称和ID
                short_name, enterprise_id = ENTERPRISE_MAP.get(company, ("", ""))
                
                num_data_rows = len(data)
                num_data_cols = len(data[0]) if data else 0
                
                # 构建新的列标题：行标题|列标题
                new_col_headers = []
                for row_idx in range(num_data_rows):  # 从第一行开始
                    row_header = row_headers[row_idx] if row_headers and row_idx < len(row_headers) else f"Row_{row_idx}"
                    # 去除所有空白字符
                    row_header = remove_all_whitespace(str(row_header))
                    for col_idx in range(num_data_cols):  # 从第一列开始
                        col_header = merged_col_headers[col_idx] if col_idx < len(merged_col_headers) else f"Col_{col_idx}"
                        # 去除所有空白字符
                        col_header = remove_all_whitespace(str(col_header))
                        new_col_header = f"{row_header}|{col_header}"
                        new_col_headers.append(new_col_header)
                
                # 构建行数据
                row_data = []
                for row_idx in range(num_data_rows):
                    for col_idx in range(num_data_cols):
                        val = None
                        v = data[row_idx][col_idx] if col_idx < len(data[row_idx]) else None
                        if v is not None and not pd.isna(v):
                            val = v
                        row_data.append(val)
                
                # 创建DataFrame
                df = pd.DataFrame([row_data], columns=new_col_headers)
                
                # 添加数据集作为第五列
                if name is not None:
                    df.insert(0, "数据集", name)
                else:
                    df.insert(0, "数据集", "")

                # 添加报表类型作为第四列
                df.insert(0, "报表类型", bracket_content)
                
                # 添加会计期间作为第三列
                if timestamp is not None:
                    df.insert(0, "会计期间", timestamp)
                else:
                    df.insert(0, "会计期间", "")
                
                # 添加企业ID作为第二列
                df.insert(0, "企业ID", enterprise_id)
                
                # 添加企业简称作为第一列
                df.insert(0, "企业简称", short_name)
                
                all_dfs.append(df)
            else:
                # 普通模式：每个sheet一个文件
                df = pd.DataFrame(data, columns=merged_col_headers)
                
                # 添加行标题列
                if row_headers:
                    repeat_count = (len(df) + len(row_headers) - 1) // len(row_headers)
                    extended_headers = row_headers * repeat_count
                    df.insert(0, "row_header", extended_headers[:len(df)])
                
                # 导出CSV
                safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
                output_path = exporter.export(df, f"{safe_sheet_name}.csv")
                output_files.append(output_path)
        
        # 合并模式：导出合并后的文件
        if merge and all_dfs:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=FutureWarning)
                merged_df = pd.concat(all_dfs, ignore_index=True)
            safe_file_name = re.sub(r'[\\/*?:"<>|]', "_", Path(file_path).stem)
            output_path = exporter.export(merged_df, f"{safe_file_name}_merged.csv")
            output_files.append(output_path)
        
        return output_files
    
    finally:
        reader.close()
