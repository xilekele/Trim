"""Excel文件读取工具"""

import warnings
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
import pandas as pd


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self._load_workbook()
    
    def _load_workbook(self):
        """加载Excel工作簿"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {self.file_path}")
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
    
    def get_sheet_names(self) -> List[str]:
        """获取所有sheet名称"""
        return self.workbook.sheetnames
    
    def get_sheet_info(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """获取指定sheet的统计信息"""
        if sheet_name is None:
            sheet_name = self.workbook.active.title
        
        ws = self.workbook[sheet_name]
        info = {
            "name": sheet_name,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "used_range": self._get_used_range(ws),
        }
        return info
    
    def _get_used_range(self, ws) -> str:
        """获取使用的单元格范围"""
        min_col = ws.min_column
        max_col = ws.max_column
        min_row = ws.min_row
        max_row = ws.max_row
        from openpyxl.utils import get_column_letter
        start = f"{get_column_letter(min_col)}{min_row}"
        end = f"{get_column_letter(max_col)}{max_row}"
        return f"{start}:{end}"
    
    def get_all_sheets_info(self) -> List[Dict[str, Any]]:
        """获取所有sheet的统计信息"""
        return [self.get_sheet_info(name) for name in self.get_sheet_names()]
    
    def read_sheet_data(
        self,
        sheet_name: Optional[str] = None,
        header_rows: Optional[Tuple[int, int]] = None,
        value_cols: Optional[Tuple[int, int]] = None,
    ) -> pd.DataFrame:
        """读取sheet数据
        
        Args:
            sheet_name: sheet名称，默认为活动sheet
            header_rows: 标题行范围 (start_row, end_row)，如 (1, 2)
            value_cols: 值列范围 (start_col, end_col)，如 (3, 10)
        
        Returns:
            DataFrame对象
        """
        if sheet_name is None:
            sheet_name = self.workbook.active.title
        
        ws = self.workbook[sheet_name]
        
        # 确定数据范围
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column
        
        # 应用标题行范围
        if header_rows:
            header_end = header_rows[1]
            min_row = header_end + 1
        
        # 应用值列范围
        if value_cols:
            min_col = value_cols[0]
            max_col = value_cols[1]
        
        # 读取数据
        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        
        if not data:
            return pd.DataFrame()
        
        # 创建DataFrame
        from openpyxl.utils import get_column_letter
        columns = [get_column_letter(i) for i in range(min_col, max_col + 1)]
        
        return pd.DataFrame(data, columns=columns)
    
    def read_sheet_with_headers(
        self,
        sheet_name: Optional[str] = None,
        header_range: Optional[Tuple[int, int, int, int]] = None,
    ) -> Tuple[pd.DataFrame, List[str]]:
        """读取sheet数据及其标题
        
        Args:
            sheet_name: sheet名称
            header_range: 标题范围 (start_row, start_col, end_row, end_col)
        
        Returns:
            (DataFrame, 合并后的标题列表)
        """
        if sheet_name is None:
            sheet_name = self.workbook.active.title
        
        ws = self.workbook[sheet_name]
        
        if header_range:
            start_row, start_col, end_row, end_col = header_range
        else:
            start_row = ws.min_row
            start_col = ws.min_column
            end_row = ws.max_row
            end_col = ws.max_column
        
        # 读取标题
        headers = []
        for row in range(start_row, end_row + 1):
            row_headers = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_headers.append(str(cell.value) if cell.value else "")
            headers.append(row_headers)
        
        # 合并标题（按列合并）
        merged_headers = []
        for col_idx in range(len(headers[0]) if headers else 0):
            parts = []
            for row_idx in range(len(headers)):
                if headers[row_idx][col_idx]:
                    parts.append(headers[row_idx][col_idx])
            merged_headers.append(" | ".join(parts))
        
        # 读取数据
        data_start_row = end_row + 1
        data = []
        for row in range(data_start_row, ws.max_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        
        columns = merged_headers
        df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)
        
        return df, merged_headers
    
    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """读取所有sheet的数据"""
        result = {}
        for sheet_name in self.get_sheet_names():
            df = self.read_sheet_data(sheet_name)
            result[sheet_name] = df
        return result
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
